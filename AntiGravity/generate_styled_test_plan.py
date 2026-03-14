import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

output_file = r"c:\Users\akank\OneDrive\Documents\All_abt_AI\TestPlanwithAI\Test_Plan_app.vwo.com.xlsx"

tabs_data = {
    "Objective": [{"Purpose": "Ensure the VWO Login Dashboard provides secure, intuitive, and eﬃcient authentication.", "Target Users": "Primary: digital marketers, product managers, UX designers, developers. Secondary: enterprise teams, CRO specialists, data analysts."}],
    "Scope": [{"Features to be tested": "Email/password auth, Remember Me, session management, password reset, 2FA, SSO (SAML/OAuth), responsive UI, accessibility (WCAG 2.1 AA), load time optimizations.", "Types of testing": "Manual functional, automated regression, performance/load, security, accessibility testing."}],
    "Inclusions": [{"Overview": "Core authentication, Enhanced UX (mobile optimization, accessibility), Enterprise features (SSO, advanced security, analytics).", "Test Objectives": "Verify authentication correctness, session management, password recovery, performance targets (sub-2-second load), and security requirements."}],
    "Exclusions": [{"Out of scope": "Features not listed in PRD (e.g., non-login platform areas, specific backend admin APIs).", "Details": "Any features outside the login dashboard boundaries."}],
    "Test Environments": [{"Operating Systems": "Windows 10, macOS, Linux, Mobile OS (for responsive design).", "Browsers / Form Factors": "Chrome, Firefox, Edge, Safari."}],
    "Defect Reporting Procedure": [{"Criteria": "Deviation from PRD requirements (functional, security, performance, accessibility).", "Steps": "Log defects in JIRA; include reproduction steps, screenshots/logs."}],
    "Test Strategy": [{"Design Techniques": "Equivalence Class Partitioning; Boundary Value Analysis; Use Case Testing; Error Guessing; Exploratory Testing.", "Procedure": "Smoke testing -> in-depth functional/regression -> cross-environment compatibility; security/pen tests and load tests per PRD."}],
    "Test Schedule": [{"Tasks": "Phase 1: Core Authentication. Phase 2: Enhanced UX. Phase 3: Enterprise Features. Includes Test plan creation, test case design, execution, and report submission.", "Details": "Timelines to be strictly aligned with project milestones."}],
    "Test Deliverables": [{"Deliverables": "Test plan, test cases/scripts, execution reports, defect reports, performance and accessibility reports.", "Description": "Documents provided at various stages of the testing lifecycle."}],
    "Entry and Exit Criteria": [{"Entry Criteria": "PRD received. Test environment and test data ready.", "Exit Criteria": "95%+ login success rate, sub-2-second load time, 0 successful brute force attacks, 100% compliance with security audit requirements, reduced support volume by 20%."}],
    "Tools": [{"Tools": "JIRA (bug tracking); screenshot/snipping tools; Word/Excel for documentation.", "Purpose": "Required for test creation, execution, and defect logging."}],
    "Risks and Mitigations": [{"Risks": "Security vulnerabilities, Performance issues under heavy load.", "Mitigations": "Regular security audits and pen testing, Auto-scaling infrastructure, Real-time performance monitoring."}]
}

# The requested Blue color for headers
blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
white_font = Font(color='FFFFFF', bold=True)

# Specific Tab Colors mapping (Hex values)
tab_colors = {
    "Objective": "FFFFFF00", # Yellow
    "Scope": "FF00B050", # Green
    "Inclusions": "FFFFC000", # Orange
    "Exclusions": "FF7F7F7F", # Grey
    "Test Environments": "FF0070C0", # Blue
    "Defect Reporting Procedure": "FFFF0000", # Red (Note: In PRD text it's "Defect Reporting")
    "Defect Reporting": "FFFF0000", # Red
    "Test Strategy": "FFFFC000", # Orange
    "Test Schedule": "FF0070C0", # Blue
    "Test Deliverables": "FF00B050", # Green
    "Entry and Exit Criteria": "FFFFFFFF", # White
    "Tools": "FFFFFF00", # Yellow
    "Risks and Mitigations": "FFFF0000" # Red
}

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, data in tabs_data.items():
        # Excel sheet names max 31 chars
        valid_sheet_name = sheet_name[:31]
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name=valid_sheet_name, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets[valid_sheet_name]
        
        # Set Tab Color based on mapping
        color_hex = tab_colors.get(sheet_name, "00000000")
        if color_hex:
            worksheet.sheet_properties.tabColor = color_hex
        
        # Set Header Formatting (Blue background, white text)
        for cell in worksheet[1]:
            cell.fill = blue_fill
            cell.font = white_font
            
        # Optional: Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Avoid too wide columns
            adjusted_width = min(max_length + 2, 80)
            worksheet.column_dimensions[col_letter].width = adjusted_width

print(f"Excel file mapped and formatted successfully at: {output_file}")
