import openpyxl
import os

excel_path = r"c:\Users\akank\OneDrive\Documents\All_abt_AI\TestPlanwithAI\AntiGravity\VWO_Login_TestPlan.xlsx"
excel_path2 = r"c:\Users\akank\OneDrive\Documents\All_abt_AI\TestPlanwithAI\VS_GithubCopilot\VWO_Login_TestPlan.xlsx"

active_path = excel_path if os.path.exists(excel_path) else excel_path2

wb = openpyxl.load_workbook(active_path, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_props = ws.sheet_properties
    tab_color = sheet_props.tabColor
    color_val = tab_color.rgb if tab_color else "None"
    
    # Check header color of A1
    cell = ws['A1']
    fill = cell.fill
    fill_color = fill.start_color.rgb if fill and fill.start_color else "None"
    font_color = cell.font.color.rgb if cell.font and cell.font.color else "None"
    
    print(f"Sheet: {sheet_name} | Tab Color: {color_val} | Header Fill: {fill_color} | Header Font: {font_color}")
    break # Just need checking the first sheet to get the theme color
