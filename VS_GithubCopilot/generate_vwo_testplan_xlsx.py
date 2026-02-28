from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

# Data derived strictly from provided PRD and user template.
VERIFIED_FACTS = [
    ("Document", "VWO Login Dashboard (app.vwo.com)"),
    ("Product Vision", "Secure, intuitive, efficient login experience connecting users to VWO platform."),
    ("Target Users", "Digital marketers; product managers; UX designers; developers; enterprise teams; CRO specialists; data analysts."),
    ("Business Objectives", "Secure access; minimize login friction; support enterprise security/compliance; facilitate onboarding."),
]

MISSING_INFO = [
    "API documentation / authentication endpoint contracts (URLs, request/response schemas, status codes).",
    "Test accounts / credentials / test data for authentication flows.",
    "Exact password complexity rules and hashing algorithm details.",
    "Specific SSO and MFA configuration details.",
    "Exact session timeout values and rate-limiting thresholds.",
    "UI mockups, exact copy for labels and error messages.",
    "Browser/OS/version matrix and device model list.",
    "Timeline / schedule and named owners for tasks.",
]

# Test plan sections mapped from user-provided template and PRD facts.
SECTIONS = {
    "Objective": [
        ("Purpose", "Ensure the VWO Login Dashboard provides secure, accessible, performant, and correct authentication and onboarding flows for the target users listed in the PRD."),
        ("Target Users", "See Verified Facts sheet."),
    ],
    "Scope": [
        ("Features to be tested", "Email/password auth; Remember Me; session management; password reset; optional MFA; SSO (SAML/OAuth) paths; validation behaviors; responsive UI; accessibility features; security behaviors; analytics emission to core dashboard."),
        ("Types of testing", "Manual functional, automated regression, performance/load, security (pen-tests), accessibility (WCAG 2.1 AA), compatibility."),
        ("Environments", "OS and browsers per user template (Windows 10, macOS, Linux; Chrome, Firefox, Edge); devices: desktop/laptop/tablet/smartphone. Exact versions not provided."),
        ("Evaluation criteria", "Defect counts; time to complete testing; user satisfaction; KPIs from PRD (page load <2s, login success rate >=95%)."),
    ],
    "Inclusions": [
        ("Overview", "Overview, purpose, scope, goals as extracted from PRD."),
        ("Test Objectives", "Verify authentication correctness, session management, password recovery, accessibility per WCAG 2.1 AA, performance targets, and security requirements per PRD."),
    ],
    "Exclusions": [
        ("Out of scope", "Features not listed in PRD (e.g., non-login platform areas). Specific backend admin APIs are out of scope unless provided).")
    ],
    "Test Environments": [
        ("Operating Systems", "Windows 10; macOS; Linux (exact versions not provided)."),
        ("Browsers", "Google Chrome; Mozilla Firefox; Microsoft Edge (versions not provided)."),
        ("Devices", "Desktop, laptop, tablet, smartphone."),
        ("Network", "Wi-Fi, cellular, wired — simulate standard connection to validate <2s load target."),
        ("Security Protocols", "HTTPS/TLS enforcement required (exact TLS versions not provided)."),
    ],
    "Defect Reporting Procedure": [
        ("Criteria", "Deviation from PRD requirements (functional, security, performance, accessibility)") ,
        ("Steps", "Log defects in tracking tool (e.g., JIRA); include reproduction steps, screenshots/logs"),
        ("Triage", "Assign severity/priority; test lead triages; developers assigned for fixes."),
    ],
    "Test Strategy": [
        ("Design Techniques", "Equivalence Class Partitioning; Boundary Value Analysis; Decision Table; State Transition; Use Case Testing; Error Guessing; Exploratory Testing."),
        ("Procedure", "Smoke testing -> in-depth functional/regression -> cross-environment compatibility; security/pen tests and load tests per PRD; accessibility checks."),
        ("Best Practices", "Shift-left testing; context-driven testing; exploratory testing; end-to-end flow testing."),
    ],
    "Test Schedule": [
        ("Tasks", "Test plan creation; test case design; test execution; report submission."),
        ("Dates", "Insufficient information to determine specific dates or timeline."),
    ],
    "Test Deliverables": [
        ("Deliverables", "Test plan; test cases/scripts; execution reports; defect reports; performance and accessibility reports."),
    ],
    "Entry and Exit Criteria": [
        ("Requirement Analysis Entry", "PRD and requirements documents received."),
        ("Requirement Analysis Exit", "Requirements clarified and test plan approved."),
        ("Test Execution Entry", "Signed-off test scenarios and test cases; application build ready." ) ,
        ("Test Execution Exit", "Test case execution completed; defects logged; regression criteria met."),
    ],
    "Tools": [
        ("Tools", "JIRA (bug tracking); screenshot/snipping tools; Word/Excel for documentation. Specific automation/perf/security tools not provided."),
    ],
    "Risks and Mitigations": [
        ("Risks", "Non-availability of resources; build/URL not working; insufficient time; security/compliance gaps if controls missing; performance issues if CDN/scaling not configured."),
        ("Mitigations", "Backup resource planning; parallel tasks; security audits and monitoring as PRD indicates.")
    ],
    "Approvals": [
        ("Approval Documents", "Test Plan; Test Scenarios; Test Cases; Test Reports.")
    ],
    "Verified Facts": [(k, v) for k, v in VERIFIED_FACTS],
    "Missing / Unknown Information": [("Item", m) for m in MISSING_INFO],
}

# Colors per sheet (openpyxl expects RGB without #)
SHEET_COLORS = {
    "Objective": "FF9999",
    "Scope": "99CCFF",
    "Inclusions": "CCFF99",
    "Exclusions": "FFCC99",
    "Test Environments": "FF99FF",
    "Defect Reporting Procedure": "FFFF99",
    "Test Strategy": "99FFCC",
    "Test Schedule": "CCCCFF",
    "Test Deliverables": "D9D9D9",
    "Entry and Exit Criteria": "B3E6FF",
    "Tools": "E6CCFF",
    "Risks and Mitigations": "FFB3B3",
    "Approvals": "CCE6FF",
    "Verified Facts": "C6E2FF",
    "Missing / Unknown Information": "FFD9B3",
}

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


wb = Workbook()
# remove default sheet
default = wb.active
wb.remove(default)

def sanitize_title(name: str) -> str:
    # Remove characters invalid in Excel sheet titles: : \/ ? * [ ] and truncate to 31 chars
    if not name:
        return "Sheet"
    cleaned = re.sub(r"[:\\\/?\*\[\]]", " ", name)
    cleaned = cleaned.strip()
    if not cleaned:
        return "Sheet"
    return cleaned[:31]


for sheet_name, rows in SECTIONS.items():
    safe_title = sanitize_title(sheet_name)
    ws = wb.create_sheet(title=safe_title)
    # set tab color if available
    color = SHEET_COLORS.get(sheet_name)
    if color:
        ws.sheet_properties.tabColor = color
    # write rows
    if isinstance(rows, list):
        r = 1
        for item in rows:
            if isinstance(item, tuple) and len(item) == 2:
                ws.cell(row=r, column=1, value=item[0])
                ws.cell(row=r, column=2, value=item[1])
                r += 1
            else:
                ws.cell(row=r, column=1, value=str(item))
                r += 1
    autofit_columns(ws)

output_path = "VWO_Login_TestPlan.xlsx"
wb.save(output_path)
print(f"Generated {output_path} in the current directory.")
